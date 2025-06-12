import os
import time
import pendulum.datetime
import pytz
import pendulum
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from pumped_hydro_dtypes import FloParamsPumpedHydro, DParams, DNode, DEdge


class DGraph():
    def __init__(self, flo_params: FloParamsPumpedHydro):
        self.params = DParams(flo_params)
        self.create_nodes()
        self.create_edges()
        self.solve_dijkstra()
        
    def create_nodes(self):
        self.nodes: Dict[int, List[DNode]] = {
            time_slice: [
                DNode(time_slice, energy_slice, self.params) 
                for energy_slice in range(self.params.energy_discretization+1)] 
            for time_slice in range(self.params.horizon+1)
        }
        self.initial_node = [n for n in self.nodes[0] if n.energy_slice==self.params.initial_energy_slice][0]
        self.min_node_energy = min(self.nodes[0], key=lambda n: n.energy).energy
        self.max_node_energy = max(self.nodes[0], key=lambda n: n.energy).energy

    def find_next_node(self, current_node: DNode, mwh_to_store: float) -> DNode:
        target_energy = current_node.energy + mwh_to_store
        return min(self.nodes[current_node.time_slice+1], key=lambda x: abs(x.energy-target_energy))
    
    def create_edges(self):
        self.edges: Dict[DNode, List[DEdge]] = {}

        for time_slice in range(self.params.horizon):

            lmp_usd_mwh = float(self.params.lmp[time_slice])
            reg_mcp_usd_mw = float(self.params.reg_mcp[time_slice])
            
            for current_node in self.nodes[time_slice]:
                self.edges[current_node] = []

                # [Generate/Discharge] as much as possible
                discharge_for_empty = current_node.energy - self.min_node_energy
                max_discharge_mw = min(self.params.generation_mw, discharge_for_empty)
                mwh_to_grid = max_discharge_mw * (1-self.params.gen_loss_percent/100) 
                mwh_to_store = -max_discharge_mw
                cost_usd = -mwh_to_grid*lmp_usd_mwh
                next_node = self.find_next_node(current_node, mwh_to_store)
                self.edges[current_node].append(DEdge(current_node, next_node, cost_usd, mwh_to_grid))

                # [Generate/Discharge] at regulation midpoint if the store is not close to empty
                if self.params.flo_params.RegulationGenerating and discharge_for_empty > self.params.generation_mw:
                    mwh_to_grid = self.params.reg_midpoint_mw 
                    mwh_to_store = -self.params.reg_midpoint_mw / (1-self.params.gen_loss_percent/100) 
                    cost_usd = -mwh_to_grid*lmp_usd_mwh - (self.params.generation_mw-self.params.reg_midpoint_mw)*reg_mcp_usd_mw
                    next_node = self.find_next_node(current_node, mwh_to_store)
                    self.edges[current_node].append(DEdge(current_node, next_node, cost_usd, mwh_to_grid))
                
                # [Pump/Charge] as much as possible
                charge_for_full = (self.max_node_energy - current_node.energy) / (1-self.params.pump_loss_percent/100)
                max_charge_mw = min(self.params.pumping_mw, charge_for_full)
                mwh_to_grid = -max_charge_mw
                mwh_to_store = max_charge_mw * (1-self.params.pump_loss_percent/100)
                cost_usd = -mwh_to_grid*lmp_usd_mwh
                next_node = self.find_next_node(current_node, mwh_to_store)
                self.edges[current_node].append(DEdge(current_node, next_node, cost_usd, mwh_to_grid))

                # [Pumping/Charge] at regulation midpoint if the store is not close to full
                if self.params.flo_params.RegulationPumping and charge_for_full < self.params.generation_mw:
                    mwh_to_grid = -self.params.reg_midpoint_mw
                    mwh_to_store = self.params.reg_midpoint_mw * (1-self.params.pump_loss_percent/100)
                    cost_usd = -mwh_to_grid*lmp_usd_mwh - (self.params.pumping_mw-self.params.reg_midpoint_mw)*reg_mcp_usd_mw
                    next_node = self.find_next_node(current_node, mwh_to_store)
                    self.edges[current_node].append(DEdge(current_node, next_node, cost_usd, mwh_to_grid))

                # Do nothing
                mwh_to_grid = 0
                mwh_to_store = 0
                cost_usd = 0
                next_node = self.find_next_node(current_node, mwh_to_store)
                self.edges[current_node].append(DEdge(current_node, next_node, cost_usd, mwh_to_grid))

            print(f"Built edges for hour {time_slice}")
    
    def solve_dijkstra(self):
        for time_slice in range(self.params.horizon-1, -1, -1):
            for node in self.nodes[time_slice]:
                best_edge = min(self.edges[node], key=lambda e: e.head.pathcost + e.cost)
                node.pathcost = best_edge.head.pathcost + best_edge.cost
                node.next_node = best_edge.head

    def export_to_excel(self):
        print("\nExporting to Excel...")
        # Along the shortest path
        generated, pumped, costs_lmp, costs_reg, costs_total = [], [], [], [], []
        node_i = self.initial_node
        while node_i.next_node is not None:
            edge_i = [e for e in self.edges[node_i] if e.tail==node_i and e.head==node_i.next_node][0]
            if edge_i.mwh_to_grid > 0:
                pump = 0
                generate = edge_i.mwh_to_grid
            else:
                pump = -edge_i.mwh_to_grid
                generate = 0
            lmp_price = self.params.lmp[node_i.time_slice]
            cost_lmp = -edge_i.mwh_to_grid * lmp_price
            cost_reg = edge_i.cost-cost_lmp if cost_lmp != edge_i.cost else 0
            pumped.append(pump)
            generated.append(generate)
            costs_lmp.append(cost_lmp)
            costs_reg.append(cost_reg)
            costs_total.append(edge_i.cost)
            node_i = node_i.next_node

        # First dataframe: the Dijkstra graph
        dijkstra_pathcosts = {}
        nodes_by_energy = sorted(self.nodes[0], key=lambda n: n.energy_slice)
        dijkstra_pathcosts['Stored MWh'] = sorted([round(n.energy,2) for n in nodes_by_energy])
        dijkstra_pathcosts['Index'] = sorted(list(range(len(nodes_by_energy))))
        dijkstra_nextnodes = dijkstra_pathcosts.copy()
        for h in range(self.params.horizon):
            dijkstra_pathcosts[h] = [round(x.pathcost,2) for x in sorted(self.nodes[h], key=lambda x: x.energy_slice)]
            dijkstra_nextnodes[h] = [x.next_node.energy_slice for x in sorted(self.nodes[h], key=lambda x: x.energy_slice)]
        dijkstra_pathcosts[self.params.horizon] = [0]*len(self.nodes[0])
        dijkstra_nextnodes[self.params.horizon] = [np.nan]*len(self.nodes[0])
        dijkstra_pathcosts_df = pd.DataFrame(dijkstra_pathcosts)
        dijkstra_nextnodes_df = pd.DataFrame(dijkstra_nextnodes)
        
        # Second dataframe: the forecasts
        start_time_unix = pendulum.datetime(self.params.year, 1, 1, 0, tz="America/New_York").timestamp() + self.params.start_hour
        start_time = datetime.fromtimestamp(start_time_unix, tz=pytz.timezone("America/New_York"))
        forecast_df = pd.DataFrame({'Prices':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.params.horizon)}})
        forecast_df.loc[0] = ['Hour'] + [start_time.strftime("%d/%m/%Y")] + [(start_time + timedelta(hours=x)).hour for x in range(self.params.horizon)]
        forecast_df.loc[1] = ['LMP'] + ['USD/MWh'] + self.params.lmp
        forecast_df.loc[2] = ['Reg MCP'] + ['USD/MW'] + self.params.reg_mcp
        
        # Third dataframe: the shortest path
        shortestpath_df = pd.DataFrame({'Shortest path':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.params.horizon+1)}})
        shortestpath_df.loc[0] = ['Pumped'] + ['MWh'] + [round(x,3) for x in pumped] + [0]
        shortestpath_df.loc[1] = ['Generated'] + ['MWh'] + [round(x,3) for x in generated] + [0]
        shortestpath_df.loc[2] = ['Profit - Total'] + ['USD'] + [-round(x,2) for x in costs_total] + [0]
        shortestpath_df.loc[4] = ['Profit - LMP'] + ['USD'] + [-round(x,2) for x in costs_lmp] + [0]
        shortestpath_df.loc[3] = ['Profit - Reg'] + ['USD'] + [-round(x,2) for x in costs_reg] + [0]

        # Fourth dataframe: the results
        results = [
            "Profit ($M)", round(-self.initial_node.pathcost/1e6,1), 
            "LMP ($M)", round(-sum(costs_lmp)/1e6,1),
            "Reg ($M)", round(-sum(costs_reg)/1e6,1),
            "Gen (MWh)", round(sum(generated)), 
            "Pump (MWh)", round(sum(pumped)),
            'PARAMETERS',
            'Year', self.params.flo_params.Year,
            'Location', self.params.flo_params.LocationId,
            'GenerationMw', self.params.flo_params.GenerationMw,
            'PumpingMw', self.params.flo_params.PumpingMw,
            'RegMidpointMw', self.params.flo_params.RegMidpointMw,
            'StoreMwh', self.params.flo_params.StoreMwh,
            'PumpLossPercent', self.params.flo_params.PumpLossPercent,
            'GenLossPercent', self.params.flo_params.GenLossPercent,
            'FloStartHr', self.params.flo_params.FloStartHr,
            'FloHours', self.params.flo_params.FloHours,
            'EnergyDiscretization', self.params.flo_params.EnergyDiscretization,
            'InitialEnergySlice', self.params.flo_params.InitialEnergySlice,
            'RegulationPumping', self.params.flo_params.RegulationPumping,
            'RegulationGenerating', self.params.flo_params.RegulationGenerating
            ]
        results_df = pd.DataFrame({'RESULTS':results})
        
        # Highlight shortest path
        highlight_positions = []
        node_i = self.initial_node
        while node_i.next_node is not None:
            highlight_positions.append((node_i.energy_slice+1+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
            node_i = node_i.next_node
        highlight_positions.append((node_i.energy_slice+1+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
        
        # Add the parameters to a seperate sheet
        parameters = self.params.flo_params.to_dict()
        parameters_df = pd.DataFrame(list(parameters.items()), columns=['Variable', 'Value'])

        # Write to Excel
        os.makedirs('results', exist_ok=True)
        file_path = os.path.join('results', f'result_{int(time.time())}.xlsx')
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False, sheet_name='Pathcost')
            results_df.to_excel(writer, index=False, sheet_name='Next node')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Pathcost')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Next node')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Pathcost')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Next node')
            dijkstra_pathcosts_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Pathcost')
            dijkstra_nextnodes_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Next node')
            parameters_df.to_excel(writer, index=False, sheet_name='Parameters')

            # Get the sheets after they've been created
            pathcost_sheet: Worksheet = writer.sheets['Pathcost']
            nextnode_sheet: Worksheet = writer.sheets['Next node']
            parameters_sheet: Worksheet = writer.sheets['Parameters']
            for row in pathcost_sheet['A1:A40']:
                for cell in row:
                    cell: Cell = cell
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in nextnode_sheet['A1:A40']:
                for cell in row:
                    cell: Cell = cell
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in parameters_sheet[f'B1:B{len(parameters_df)+1}']:
                for cell in row:
                    cell: Cell = cell
                    cell.alignment = Alignment(horizontal='right')
            pathcost_sheet.column_dimensions['A'].width = 17.5
            pathcost_sheet.column_dimensions['B'].width = 15
            pathcost_sheet.column_dimensions['C'].width = 15
            nextnode_sheet.column_dimensions['A'].width = 17.5
            nextnode_sheet.column_dimensions['B'].width = 15
            nextnode_sheet.column_dimensions['C'].width = 15
            parameters_sheet.column_dimensions['A'].width = 40
            parameters_sheet.column_dimensions['B'].width = 70
            pathcost_sheet.freeze_panes = 'D12'
            nextnode_sheet.freeze_panes = 'D12'

            # Highlight shortest path
            highlight_fill = PatternFill(start_color='72ba93', end_color='72ba93', fill_type='solid')
            highlight_fill_2 = PatternFill(start_color='e8b38e', end_color='e8b38e', fill_type='solid')
            for row in range(len(forecast_df)+len(shortestpath_df)+2+1):
                pathcost_sheet.cell(row=row+1, column=1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=1).fill = highlight_fill
            for row in range(len(forecast_df)+len(shortestpath_df)+2+1,len(forecast_df)+len(shortestpath_df)+2+1+29):
                pathcost_sheet.cell(row=row+1, column=1).fill = highlight_fill_2
                nextnode_sheet.cell(row=row+1, column=1).fill = highlight_fill_2
            for row, col in highlight_positions:
                pathcost_sheet.cell(row=row+1, column=col+1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=col+1).fill = highlight_fill


if __name__ == "__main__":

    df = pd.read_csv('price_data.csv')
    df['lmp'] = df['lmp'].str.replace(',', '').astype(float)
    df['regmcp'] = df['regmcp'].str.replace(',', '').astype(float)
    lmp = list(df['lmp'])
    reg = list(df['regmcp'])

    print("\nHi George")
    year = input("Year (leave blank for 2024): ")
    year = 2024 if year == "" else int(year)
    flo_start_hr = input("FloStartHr (leave blank for 1): ")
    flo_start_hr = 1 if flo_start_hr == "" else int(flo_start_hr)
    flo_hours = input("FloHours (leave blank for 8760): ")
    flo_hours = 8760 if flo_hours == "" else int(flo_hours)
    store_mwh = input("StoreMwh (leave blank for 2800): ")
    store_mwh = 2800 if store_mwh == "" else int(store_mwh)
    generation_mw = input("GenerationMw (leave blank for 500): ")
    generation_mw = 500 if generation_mw == "" else int(generation_mw)
    pumping_mw = input("PumpingMw (leave blank for 500): ")
    pumping_mw = 500 if pumping_mw == "" else int(pumping_mw)
    regulation_pumping = input("RegulationPumping ('y' or 'n', leave blank for 'y'): ")
    regulation_pumping = True if regulation_pumping!="n" else False
    regulation_generating = input("RegulationGenerating ('y' or 'n', leave blank for 'y'): ")
    regulation_generating = True if regulation_generating!="n" else False
    reg_midpoint_mw = input("RegMidpointMw (leave blank for 350): ")
    reg_midpoint_mw = 350 if reg_midpoint_mw == "" else int(reg_midpoint_mw)
    pump_loss_percent = input("PumpLossPercent (leave blank for 10): ")
    pump_loss_percent = 10 if pump_loss_percent == "" else int(pump_loss_percent)
    gen_loss_percent = input("GenLossPercent (leave blank for 10): ")
    gen_loss_percent = 10 if gen_loss_percent == "" else int(gen_loss_percent)
    location_id = input("LocationId (leave blank for 4001): ")
    location_id = 4001 if location_id == "" else int(location_id)
    energy_discretization = input("EnergyDiscretization (leave blank for 100): ")
    energy_discretization = 100 if energy_discretization == "" else int(energy_discretization)
    initial_energy_slice = input("InitialEnergySlice (leave blank for 50): ")
    initial_energy_slice = 50 if initial_energy_slice == "" else int(initial_energy_slice)

    flo_params = FloParamsPumpedHydro(
        Year=int(year),
        FloStartHr=int(flo_start_hr),
        FloHours=int(flo_hours),
        StoreMwh=int(store_mwh),
        GenerationMw=int(generation_mw),
        PumpingMw=int(pumping_mw),
        RegMidpointMw=int(reg_midpoint_mw),
        PumpLossPercent=int(pump_loss_percent),
        GenLossPercent=int(gen_loss_percent),
        LocationId=int(location_id),
        Lmp=lmp,
        RegMcp=reg,
        EnergyDiscretization=int(energy_discretization),
        InitialEnergySlice=int(initial_energy_slice),
        RegulationPumping=regulation_pumping,
        RegulationGenerating=regulation_generating
    )
    dgraph = DGraph(flo_params)
    dgraph.export_to_excel()
    print("\nDone!")